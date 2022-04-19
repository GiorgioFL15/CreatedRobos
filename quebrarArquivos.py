from fileinput import filename
from openpyxl import load_workbook
from openpyxl import Workbook
import os

nome_arquivo = "/home/fontes-1064/workspace/criandorobos/ListaDeDolar/Quebrar.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

sheet_selecionada = planilha_aberta['Dados']

criandoNovoArquivoExcel = Workbook()

nomeNovo = ""
totalLinha = len(sheet_selecionada['A']) + 1

for linha in range(2, len(sheet_selecionada['A']) + 1):

    nomeAtual = sheet_selecionada['A%s % linha'].value

    if nomeNovo == nomeAtual:

        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A']) +1
        celulaColunaA = "A" + str(linhaSheetQuebra)
        celulaColunaB = "B" + str(linhaSheetQuebra)
        celulaColunaC = "C" + str(linhaSheetQuebra)

        selecionaSheetVendasNovaPlanilha[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaC] = sheet_selecionada['C%s' % linha].value

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)

    else:

        nomeNovo = sheet_selecionada['A%s % linha'].value

        nova_planilha = criandoNovoArquivoExcel.active

        nova_planilha.title = "Vendas"

        caminhoNovaPlanilha = "/home/fontes-1064/workspace/criandorobos/ListaDeDolar/"+ sheet_selecionada['A%s % linha'].value +" .xlsx"

        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']

        selecionaSheetVendasNovaPlanilha['A1'] = "Vendedor"
        selecionaSheetVendasNovaPlanilha['B1'] = "Produtos"
        selecionaSheetVendasNovaPlanilha['C1'] = "Vendas"

        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value

        selecionaSheetVendasNovaPlanilha.delete_rows(3, 100)

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)



#------------------ Disparando emails
    import win32com.client as win32

    outlook = win32.Dispatch('outlook.application')

    nome_arquivo = "/home/fontes-1064/workspace/criandorobos/ListaDeDolar/Quebrar.xlsx"
    planilha_aberta = load_workbook(filename=nome_arquivo)

    sheet_selecionada = planilha_aberta['Dados']

    for linha in range(2, len(sheet_selecionada['A']) +1):
        nome = sheet_selecionada['A%s % linha'].value
        nomeCompleto = sheet_selecionada['B%s % linha'].value
        email = sheet_selecionada['C%s % linha'].value

        emailOutlook = outlook.CreateItem(0)

        emailOutlook.To = email
        emailOutlook.Subject = "Lista de vendas" + nomeCompleto
        emailOutlook.HTMLBody = f"""
        <p>Boa noite <b>{nome}</b></p>
        <p>Segue o relatório com suas vendas.</p>
        <p>Atenciosamente.</p>
        """

        anexoEmail = "/home/fontes-1064/workspace/criandorobos/ListaDeDolar/"+ nomeCompleto + ".xlsx"

        emailOutlook.save() #save = Cria e salva o email, Send() - Enviar o email